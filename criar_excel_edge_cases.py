#!/usr/bin/env python3
"""
Script para criar arquivos Excel de teste para edge cases
"""

from openpyxl import Workbook

# Teste 1: Colunas em ordem invertida
print("Criando teste 1: Colunas em ordem invertida...")
wb1 = Workbook()
ws1 = wb1.active
ws1.title = "Dados"

dados1 = [
    ['CODIGOBARRAS', 'IMBLOJA'],  # Ordem invertida
    ['7896050201756', '0001002154'],
    ['7898080070050', '0001002154'],
]

for row in dados1:
    ws1.append(row)

wb1.save('teste_ordem_invertida.xlsx')
print("✓ teste_ordem_invertida.xlsx criado")

# Teste 2: Nomes de colunas em lowercase
print("\nCriando teste 2: Colunas em lowercase...")
wb2 = Workbook()
ws2 = wb2.active
ws2.title = "Dados"

dados2 = [
    ['imbloja', 'codigobarras'],  # Lowercase
    ['0001002154', '7896050201756'],
    ['0001002154', '7898080070050'],
]

for row in dados2:
    ws2.append(row)

wb2.save('teste_lowercase.xlsx')
print("✓ teste_lowercase.xlsx criado")

# Teste 3: Arquivo com linhas vazias
print("\nCriando teste 3: Arquivo com linhas vazias...")
wb3 = Workbook()
ws3 = wb3.active
ws3.title = "Dados"

dados3 = [
    ['IMBLOJA', 'CODIGOBARRAS'],
    ['0001002154', '7896050201756'],
    ['', ''],  # Linha vazia
    ['0001002154', '7898080070050'],
    ['', ''],  # Linha vazia
]

for row in dados3:
    ws3.append(row)

wb3.save('teste_linhas_vazias.xlsx')
print("✓ teste_linhas_vazias.xlsx criado")

# Teste 4: Mixed case
print("\nCriando teste 4: Mixed case...")
wb4 = Workbook()
ws4 = wb4.active
ws4.title = "Dados"

dados4 = [
    ['ImBLoJa', 'CoDiGoBarRaS'],  # Mixed case
    ['0001002154', '7896050201756'],
    ['0001002154', '7898080070050'],
]

for row in dados4:
    ws4.append(row)

wb4.save('teste_mixed_case.xlsx')
print("✓ teste_mixed_case.xlsx criado")

print("\n✅ Todos os arquivos de teste criados com sucesso!")
